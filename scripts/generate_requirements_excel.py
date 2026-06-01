#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""生成满改平台需求清单 Excel 文件"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "需求清单.xlsx"

# ============== 样式定义 ==============
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=12, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUBHEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
MODULE_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
MODULE_FONT = Font(name="微软雅黑", size=11, bold=True, color="000000")
NORMAL_FONT = Font(name="微软雅黑", size=10, color="000000")
BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)

def set_header_style(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def set_subheader_style(cell):
    cell.fill = SUBHEADER_FILL
    cell.font = SUBHEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = BORDER

def set_module_style(cell):
    cell.fill = MODULE_FILL
    cell.font = MODULE_FONT
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border = BORDER

def set_cell_style(cell, align="left"):
    cell.font = NORMAL_FONT
    cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=True)
    cell.border = BORDER

def auto_width(ws, min_width=10, max_width=50):
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val = str(cell.value) if cell.value else ""
                lines = val.split("\n")
                for line in lines:
                    max_length = max(max_length, len(line))
            except:
                pass
        adjusted = min(max(min_width, max_length + 2), max_width)
        ws.column_dimensions[col_letter].width = adjusted

# ============== 数据 ==============

OVERVIEW_ROWS = [
    ["项目", "说明"],
    ["项目名称", "满改 — 汽车改装/周边售卖/服务预约平台"],
    ["核心模式", "内容（社区/资讯）引流 → 商城转化 + 服务履约"],
    ["参考风格", "理想汽车 APP（社区+商城+服务+我的）"],
    ["数据方案", "全部 Mock 填充，不接入真实接口/数据库"],
    ["支付方案", "预留支付宝/微信支付入口，Mock 支付状态"],
    ["物流方案", "平台手动录入单号（一期）；预留 ERP 自动同步（二期）"],
    ["发票方案", "用户申请 → 平台后台处理 → 上传 PDF/邮箱回传"],
    ["自提口径", "个人自行取货，非寄到门店后再提货"],
    ["服务价口径", "商品价格与改装服务费分开计算，下单时明确说明"],
]

USER_APP_ROWS = [
    ["模块", "功能点", "说明", "状态", "备注"],
    # 首页/社区
    ["首页/社区", "内容 Feed 流", "官方资讯+精选案例+用户帖子混合流，首屏呈现", "待开发", "论坛权重提升为 Tab1"],
    ["", "分类筛选", "官方/精选/关注动态/最热/最新", "待开发", ""],
    ["", "场景板块", "动力、外观、内饰、底盘等分类入口", "待开发", ""],
    ["", "车型圈子", "新能源、燃油、房车、性能车", "待开发", ""],
    ["", "搜索", "内容/商品/门店搜索", "待开发", ""],
    ["", "商品快捷入口", "内容流中嵌入商品链接（带图/视频）", "待开发", ""],
    ["", "立即咨询", "案例下方挂改装咨询问卷，收集预算/喜好", "待开发", ""],
    # 商城
    ["商城", "首屏大图/视频", "品牌视觉展示，轮播或沉浸式大图，先视觉冲击再选购", "待开发", "不要一上来就分类"],
    ["", "签约品牌 Logo 滚动", "所有合作品牌横向滚动展示", "待开发", ""],
    ["", "商品分类", "轮毂、车衣、套件等功能板块分类", "待开发", ""],
    ["", "商品列表", "大图卡片流，非密集表格", "待开发", ""],
    ["", "商品详情", "规格、适配车型、品牌、价格、图文详情、收藏", "已有", "user-product-detail.html"],
    ["", "购物车", "增删改、结算入口", "待开发", ""],
    ["", "收藏", "收藏商品+降价通知（预留），后台统计收藏量", "待开发", ""],
    ["", "下单页", "自提/指定服务商、支付方式选择、地址填写", "已有", "user-order-create.html"],
    # 爱车
    ["爱车", "车辆 3D 模型", "可旋转车身（原型可用图片/Slider 模拟）", "待开发", ""],
    ["", "车衣颜色选择", "实时预览改色效果", "待开发", ""],
    ["", "轮毂颜色/款式选择", "实时预览换装效果", "待开发", ""],
    ["", "基础车辆信息", "品牌、车型、改装历史", "待开发", ""],
    ["", "推荐改装方案", "根据车型推荐适配轮毂+车衣组合", "待开发", ""],
    ["", "关联商品下单", "方案确认后一键生成订单（跳转下单）", "待开发", "核心转化路径"],
    # 论坛
    ["论坛", "帖子列表", "封面+标题+作者+互动数据", "待开发", ""],
    ["", "帖子详情", "视频/图片/文字混排、商品链接挂载、评论", "已有", "user-topic-detail.html"],
    ["", "发帖页", "富文本编辑器，支持视频/图片上传", "已有", "user-topic-create.html"],
    ["", "评论/回复", "评论列表、回复、点赞", "待完善", ""],
    ["", "案例详情", "精品改装案例展示", "已有", "user-case-detail.html"],
    ["", "资讯详情", "官方文章内容", "已有", "user-news-detail.html"],
    # 我的
    ["我的", "个人资料", "头像、昵称、绑定车辆", "待开发", ""],
    ["", "历史订单", "商品订单+服务订单，状态跟踪", "待开发", ""],
    ["", "购物车入口", "与商城购物车联动", "待开发", ""],
    ["", "消息中心", "系统通知、与服务商的聊天", "待开发", ""],
    ["", "开票管理", "申请开票、开票历史、发票下载（PDF/邮箱）", "待开发", ""],
    ["", "金融入口", "授信额度展示、申请入口（仅入口+记录）", "待开发", ""],
    ["", "我的发布", "帖子/案例管理", "待开发", ""],
    ["", "关注/粉丝", "社交关系", "待开发", ""],
    ["", "积分/钱包", "积分展示（预留）", "待开发", ""],
    ["", "设置", "账号安全、隐私、关于", "待开发", ""],
]

PLATFORM_ROWS = [
    ["模块", "功能点", "说明", "状态", "备注"],
    # 首页
    ["首页/Dashboard", "平台概况统计", "订单量、用户数、营收概览", "待开发", ""],
    ["", "快捷入口", "待审核服务商、待处理订单、待审帖子", "待开发", ""],
    ["", "核心门店排行", "施工量、好评率排名", "待开发", ""],
    ["", "告警信息", "异常订单、违规内容提示", "待开发", ""],
    # 服务商管理
    ["服务商管理", "入驻审核", "资质证照、门店信息审核", "待开发", ""],
    ["", "服务商列表", "经营状态、接单状态控制", "待开发", ""],
    ["", "服务商账号", "开通子账号（可挂链接的官方账号）", "待开发", ""],
    ["", "服务统计", "指定订单 vs 推荐订单金额拆分", "待开发", ""],
    # 用户管理
    ["用户管理", "用户列表", "注册用户信息、操作记录", "待开发", ""],
    ["", "绑定车辆", "用户车辆信息、改装历史", "待开发", ""],
    ["", "禁言/封禁", "7天/30天/半年/永久禁言", "待开发", ""],
    # 商品
    ["商品管理", "品牌管理", "签约品牌维护、授权文件上传", "待开发", ""],
    ["", "商品分类", "大类管理", "待开发", ""],
    ["", "商品详情", "规格、价格、库存、适配车型、图文编辑", "待开发", ""],
    ["", "收藏统计", "各商品收藏量（用于跟厂家谈活动）", "待开发", ""],
    # 车型素材
    ["车型与素材", "车型管理", "品牌、车系、年款维护", "待开发", ""],
    ["", "车衣素材", "颜色、贴图与车型适配", "待开发", ""],
    ["", "轮毂素材", "款式、颜色与车型适配", "待开发", ""],
    # 服务项目
    ["服务项目", "服务项目模板", "外观提升、动力改装等项目", "待开发", ""],
    ["", "基准价设置", "平台建议工时费", "待开发", ""],
    ["", "浮动比例", "允许服务商在±12%范围内定价", "待开发", ""],
    # 订单
    ["订单管理", "商品订单", "待付款/待发货/已发货/已完成，查看实付金额", "待开发", ""],
    ["", "服务订单", "待指派/待接单/施工中/待验收/已完成", "待开发", ""],
    ["", "一键派单", "根据用户位置推荐周边门店指派", "待开发", ""],
    ["", "用户指定服务商", "用户指定后平台不可更改", "待开发", ""],
    ["", "订单详情", "支付方式、服务商、商品清单", "待开发", ""],
    # 物流
    ["物流管理", "手动发货", "选择物流、填写运单号、备注", "待开发", ""],
    ["", "物流记录", "平台内物流状态展示", "待开发", ""],
    ["", "ERP 对接预留", "后续自动同步物流单号", "待开发", "二期"],
    # 案例论坛
    ["案例与论坛", "案例新增/编辑", "富文本、图片、视频，设置展示/隐藏", "待开发", ""],
    ["", "案例审核", "服务商提交案例的平台审核", "待开发", ""],
    ["", "案例置顶/推荐", "控制首页展示", "待开发", ""],
    ["", "版面管理", "论坛板块增删改", "待开发", ""],
    ["", "话题管理", "话题新建、帖子关联", "待开发", ""],
    ["", "版主申请审核", "服务商申请版主", "待开发", ""],
    ["", "内容审核", "帖子/评论删除、违规处理", "待开发", ""],
    # 活动
    ["活动与营销", "优惠券发放", "类型、数量、折扣上限、适用商品/品牌", "待开发", ""],
    ["", "活动板块", "限时折扣、满减、618/818 等活动配置", "待开发", ""],
    ["", "核销统计", "领取量、使用量、核销比例", "待开发", ""],
    # 发票结算
    ["发票与结算", "发票申请列表", "用户申请记录，专票/普票", "待开发", ""],
    ["", "发票处理", "上传 PDF、回传、标记已开票", "待开发", ""],
    ["", "订单流水", "实付金额、优惠券抵扣、平台到账记录", "待开发", ""],
    # 系统
    ["系统配置", "自动验收", "24h/3天/1周 自动确认配置", "待开发", ""],
    ["", "消息模板", "短信/推送模板编辑", "待开发", ""],
    ["", "访客统计", "未注册用户浏览记录", "待开发", ""],
    ["", "账号权限", "角色管理、挂链接权限控制", "待开发", ""],
]

PROVIDER_ROWS = [
    ["端", "模块", "功能点", "说明", "状态"],
    ["服务商 Web", "门店展示", "门店展示页", "品牌形象页，案例、基本信息、施工数据", "待开发"],
    ["", "案例维护", "案例维护", "富文本编辑器发布案例，提交平台审核", "待开发"],
    ["", "资质更新", "资质更新", "门店资料、证照维护", "待开发"],
    ["服务商 APP", "首页", "首页统计", "今日到店/完工数、营收统计", "待开发"],
    ["", "订单", "订单列表", "待接单/施工中/待验收/已完成", "待开发"],
    ["", "", "接单/拒单", "拒单需填原因，平台中转不直接通知用户", "待开发"],
    ["", "", "施工排期", "选择可施工时间（以用户时间为准）", "待开发"],
    ["", "", "提交完工", "上传车辆完工照片、说明", "待开发"],
    ["", "运营", "服务定价", "从平台项目库选择并设定门店报价", "待开发"],
    ["", "", "采购记录", "在平台采购商品的记录（To B 预留）", "待开发"],
    ["", "论坛", "论坛管理", "版主权限：帖子/评论删除", "待开发"],
    ["", "消息", "IM", "与用户、平台的聊天，聊天记录平台可抓取", "待开发"],
    ["", "我的", "门店资料", "信息更新、联系平台", "待开发"],
    ["", "", "营业统计", "今日营收、施工量", "待开发"],
]

BRAND_ROWS = [
    ["模块", "功能点", "说明", "状态", "备注"],
    ["品牌方 Web", "订单查看", "只能看到本品牌相关订单", "待开发", ""],
    ["", "发货确认", "填写物流单号、确认发货", "待开发", "核心动作"],
    ["", "库存维护", "库存数量同步", "待开发", "二期预留"],
]

PROCESS_ROWS = [
    ["流程名称", "步骤", "说明", "关键规则"],
    ["商品购买", "1. 浏览", "内容流/商城 → 商品详情", ""],
    ["", "2. 加购", "加入购物车 / 立即购买", ""],
    ["", "3. 确认订单", "选择配送方式（自提/指定服务商）", "自提=个人自行取货"],
    ["", "4. 支付", "选择支付方式（预留支付宝/微信）", "Mock 支付状态"],
    ["", "5. 发货", "平台/品牌方发货，填写运单号", ""],
    ["", "6. 物流", "物流跟踪", "一期手动录入，二期 ERP 同步"],
    ["", "7. 签收", "用户确认收货", ""],
    ["改装服务", "1. 选购方案", "爱车页定制或商城选购含服务的商品", "商品价格与改装服务费分开"],
    ["", "2. 下单", "选择服务门店或平台指派", "用户指定后平台不可更改"],
    ["", "3. 派单", "平台指派/用户指定服务商", ""],
    ["", "4. 接单", "服务商接单或拒单", "拒单需填原因，平台中转处理"],
    ["", "5. 施工", "到店施工，服务商提交完工", ""],
    ["", "6. 验收", "用户验收并评分", "24h/3天/1周 自动验收可配"],
    ["", "7. 完成", "订单完成", ""],
    ["内容发布", "1. 编辑", "用户/服务商编辑帖子/案例", ""],
    ["", "2. 提交", "提交平台审核", ""],
    ["", "3. 审核", "平台审核通过/驳回", ""],
    ["", "4. 展示", "正常展示或首页推荐/置顶", "官方账号可挂商品链接"],
    ["发票申请", "1. 申请", "订单完成后用户申请开票", "专票/普票"],
    ["", "2. 处理", "平台后台处理", ""],
    ["", "3. 回传", "上传 PDF、发送邮箱", ""],
]

FILE_STATUS_ROWS = [
    ["文件路径", "文件大小", "状态", "说明"],
    ["index.html", "-", "待完善", "项目入口"],
    ["user-app.html", "698 B", "占位", "用户端首页框架，需重写"],
    ["user-product-detail.html", "7,221 B", "有基础", "商品详情，需按需求完善"],
    ["user-order-create.html", "17,294 B", "较全", "下单页，需补充支付方式、自提/服务商选择"],
    ["user-topic-detail.html", "10,354 B", "有基础", "帖子详情，需完善评论、商品链接展示"],
    ["user-topic-create.html", "6,501 B", "有基础", "发帖页，需完善富文本/视频上传"],
    ["user-case-detail.html", "4,930 B", "有基础", "案例详情，需完善"],
    ["user-news-detail.html", "4,366 B", "有基础", "资讯详情，需完善"],
    ["platform-web.html", "1,550 B", "占位", "平台管理端，待开发"],
    ["provider-web.html", "1,040 B", "占位", "服务商 Web，待开发"],
    ["provider-app.html", "642 B", "占位", "服务商 APP，待开发"],
    ["admin-app.html", "642 B", "占位", "平台 APP，待开发"],
    ["brand-web.html", "1,155 B", "占位", "品牌方 Web，待开发"],
]

def build_sheet(wb, title, rows, freeze_panes="A2"):
    ws = wb.create_sheet(title=title)
    for r_idx, row in enumerate(rows, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                set_header_style(cell)
            elif len(row) > 1 and row[0] and row[0] != "" and c_idx == 1 and r_idx > 1:
                # 模块名合并行首列，使用模块样式
                if rows[r_idx-1][0] != rows[r_idx-2][0] if r_idx > 1 else True:
                    pass
            set_cell_style(cell)
    
    # 特殊处理：合并同类模块单元格（仅针对有模块列的 sheet）
    merge_module_cells(ws, rows)
    
    ws.freeze_panes = freeze_panes
    auto_width(ws)
    return ws

def merge_module_cells(ws, rows):
    """对第一列相同的连续行，合并单元格并设置模块样式"""
    if len(rows) <= 1:
        return
    start_row = 2
    current_module = rows[1][0] if len(rows[1]) > 0 else None
    for r_idx in range(2, len(rows) + 1):
        module = rows[r_idx - 1][0] if len(rows[r_idx - 1]) > 0 else None
        if module != current_module or r_idx == len(rows):
            end_row = r_idx - 1 if module == current_module else r_idx - 1
            if current_module and start_row <= end_row and end_row > start_row:
                ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                cell = ws.cell(row=start_row, column=1)
                set_module_style(cell)
            current_module = module
            start_row = r_idx
    # 处理最后一组
    if current_module and start_row <= len(rows):
        ws.merge_cells(start_row=start_row, start_column=1, end_row=len(rows), end_column=1)
        cell = ws.cell(row=start_row, column=1)
        set_module_style(cell)

def build_overview_sheet(wb):
    ws = wb.create_sheet(title="项目概览", index=0)
    for r_idx, row in enumerate(OVERVIEW_ROWS, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                set_header_style(cell)
            else:
                set_cell_style(cell)
                if c_idx == 1:
                    cell.font = Font(name="微软雅黑", size=10, bold=True, color="000000")
    ws.freeze_panes = "A2"
    auto_width(ws)
    return ws

def build_file_status_sheet(wb):
    ws = wb.create_sheet(title="现有文件状态")
    for r_idx, row in enumerate(FILE_STATUS_ROWS, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                set_header_style(cell)
            else:
                set_cell_style(cell)
                # 状态列颜色
                if c_idx == 3:
                    val = str(value)
                    if "占位" in val:
                        cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                        cell.font = Font(name="微软雅黑", size=10, color="9C0006")
                    elif "已有" in val or "较全" in val:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                        cell.font = Font(name="微软雅黑", size=10, color="006100")
                    elif "待完善" in val or "有基础" in val:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                        cell.font = Font(name="微软雅黑", size=10, color="9C5700")
    ws.freeze_panes = "A2"
    auto_width(ws)
    return ws

def build_process_sheet(wb):
    ws = wb.create_sheet(title="核心业务流程")
    for r_idx, row in enumerate(PROCESS_ROWS, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                set_header_style(cell)
            else:
                set_cell_style(cell)
    
    # 合并流程名列
    if len(PROCESS_ROWS) > 1:
        start_row = 2
        current = PROCESS_ROWS[1][0]
        for r_idx in range(2, len(PROCESS_ROWS) + 1):
            val = PROCESS_ROWS[r_idx - 1][0]
            if val != current or r_idx == len(PROCESS_ROWS):
                end_row = r_idx - 1 if val == current else r_idx - 1
                if current and start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    cell = ws.cell(row=start_row, column=1)
                    set_module_style(cell)
                current = val
                start_row = r_idx
        if current and start_row <= len(PROCESS_ROWS):
            if start_row < len(PROCESS_ROWS):
                ws.merge_cells(start_row=start_row, start_column=1, end_row=len(PROCESS_ROWS), end_column=1)
                cell = ws.cell(row=start_row, column=1)
                set_module_style(cell)
    
    ws.freeze_panes = "A2"
    auto_width(ws)
    return ws

def build_provider_sheet(wb):
    ws = wb.create_sheet(title="服务商端")
    for r_idx, row in enumerate(PROVIDER_ROWS, 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if r_idx == 1:
                set_header_style(cell)
            else:
                set_cell_style(cell)
    
    # 合并端列
    if len(PROVIDER_ROWS) > 1:
        start_row = 2
        current = PROVIDER_ROWS[1][0]
        for r_idx in range(2, len(PROVIDER_ROWS) + 1):
            val = PROVIDER_ROWS[r_idx - 1][0]
            if val != current or r_idx == len(PROVIDER_ROWS):
                end_row = r_idx - 1 if val == current else r_idx - 1
                if current and start_row < end_row:
                    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row, end_column=1)
                    cell = ws.cell(row=start_row, column=1)
                    set_module_style(cell)
                current = val
                start_row = r_idx
        if current and start_row <= len(PROVIDER_ROWS):
            if start_row < len(PROVIDER_ROWS):
                ws.merge_cells(start_row=start_row, start_column=1, end_row=len(PROVIDER_ROWS), end_column=1)
                cell = ws.cell(row=start_row, column=1)
                set_module_style(cell)
    
    # 合并模块列（端内合并）
    # 简化处理：只在同一端内合并模块列
    ws.freeze_panes = "A2"
    auto_width(ws)
    return ws

def main():
    wb = openpyxl.Workbook()
    
    # 删除默认 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    
    build_overview_sheet(wb)
    build_sheet(wb, "用户端APP", USER_APP_ROWS)
    build_sheet(wb, "平台管理端", PLATFORM_ROWS)
    build_provider_sheet(wb)
    build_sheet(wb, "品牌方端", BRAND_ROWS)
    build_process_sheet(wb)
    build_file_status_sheet(wb)
    
    wb.save(OUTPUT_PATH)
    print(f"已生成: {OUTPUT_PATH}")
    print(f"包含 {len(wb.sheetnames)} 个 Sheet: {', '.join(wb.sheetnames)}")

if __name__ == "__main__":
    main()
